"""图片 → Excel — 使用 OCR 识别图片中的表格并导出为 Excel"""
import os
import threading
from typing import Optional

from PIL import Image
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from src.core.converter_base import BaseConverter
from src.core.converter_registry import register
from src.core.models import ConversionType
from src.utils.image_utils import preprocess_for_ocr


@register(ConversionType.IMAGE_TO_EXCEL)
class ImageToExcelConverter(BaseConverter):
    """使用 OCR 识别图片中的表格并导出为 Excel"""

    def convert(
        self,
        input_path: str,
        output_dir: str,
        options: Optional[dict] = None,
        cancel_event: Optional[threading.Event] = None,
    ) -> str:
        options = options or {}
        lang = options.get("language", "chi_sim+eng")
        preprocess = options.get("preprocess", True)

        # 加载并预处理图片
        image = Image.open(input_path)
        if preprocess:
            image = preprocess_for_ocr(image)

        self._check_cancelled(cancel_event)

        # OCR 识别
        try:
            ocr_data = pytesseract.image_to_data(
                image, lang=lang, output_type=pytesseract.Output.DICT
            )
        except pytesseract.TesseractNotFoundError:
            raise RuntimeError(
                "未找到 Tesseract OCR 引擎。请确保已安装 Tesseract OCR，"
                "或联系开发者获取包含 OCR 引擎的完整版本。"
            )

        self._check_cancelled(cancel_event)

        # 将 OCR 数据组织为行列结构
        row_data = self._organize_into_table(ocr_data)

        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "识别结果"

        # 写入数据
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(name="微软雅黑", bold=True, size=11)
        body_font = Font(name="微软雅黑", size=10)

        for i, row in enumerate(row_data):
            for j, cell_value in enumerate(row):
                cell = ws.cell(row=i + 1, column=j + 1, value=cell_value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if i == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = body_font

        # 自适应列宽
        if row_data:
            for col_idx in range(1, len(row_data[0]) + 1):
                max_length = 0
                for row_idx in range(1, len(row_data) + 1):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                    char_len = sum(2 if ord(c) > 127 else 1 for c in cell_value[:50])
                    max_length = max(max_length, char_len)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 60)

        image.close()

        output_path = self._make_output_path(input_path, output_dir, ".xlsx")
        wb.save(output_path)
        return output_path

    def _organize_into_table(self, ocr_data: dict) -> list[list[str]]:
        """将 OCR 数据组织为二维表格"""
        # 收集有效文本块
        blocks = []
        for i in range(len(ocr_data["text"])):
            text = ocr_data["text"][i].strip()
            if not text:
                continue
            conf = int(ocr_data["conf"][i]) if ocr_data["conf"][i] != "-1" else 0
            if conf < 30:
                continue

            blocks.append({
                "x": ocr_data["left"][i],
                "y": ocr_data["top"][i],
                "w": ocr_data["width"][i],
                "h": ocr_data["height"][i],
                "text": text,
            })

        if not blocks:
            return [["未识别到文字内容"]]

        # 按 y 坐标排序并分组为行
        blocks.sort(key=lambda b: (b["y"], b["x"]))

        # 使用自适应行分组
        rows = []
        current_row = [blocks[0]]
        current_y = blocks[0]["y"]
        avg_height = sum(b["h"] for b in blocks) / len(blocks)

        for block in blocks[1:]:
            if block["y"] - current_y > avg_height * 0.6:
                # 新行
                current_row.sort(key=lambda b: b["x"])
                rows.append(current_row)
                current_row = [block]
                current_y = block["y"]
            else:
                current_row.append(block)

        # 最后一行
        current_row.sort(key=lambda b: b["x"])
        rows.append(current_row)

        # 确定最大列数
        max_cols = max(len(row) for row in rows) if rows else 1

        # 构建二维数组
        result = []
        for row in rows:
            result.append([b["text"] for b in row])
            # 补齐不足的列
            while len(result[-1]) < max_cols:
                result[-1].append("")

        return result
