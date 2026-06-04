"""PDF → Excel — 提取 PDF 中表格到 Excel"""
import os
import threading
from typing import Optional

import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from src.core.converter_base import BaseConverter
from src.core.converter_registry import register
from src.core.models import ConversionType


@register(ConversionType.PDF_TO_EXCEL)
class PdfToExcelConverter(BaseConverter):
    """将 PDF 中的表格提取到 Excel"""

    def convert(
        self,
        input_path: str,
        output_dir: str,
        options: Optional[dict] = None,
        cancel_event: Optional[threading.Event] = None,
    ) -> str:
        options = options or {}
        preserve_formatting = options.get("preserve_formatting", True)

        doc = fitz.open(input_path)
        wb = Workbook()

        # 移除默认工作表
        wb.remove(wb.active)

        table_count = 0
        text_lines = []

        for page_num, page in enumerate(doc):
            self._check_cancelled(cancel_event)

            # 检测表格
            try:
                tables = page.find_tables()
                if tables and tables.tables:
                    for table in tables.tables:
                        data = table.extract()
                        if data and len(data) > 0:
                            table_count += 1
                            ws = wb.create_sheet(title=f"表格{table_count}")
                            self._write_table_to_sheet(ws, data, preserve_formatting)
            except Exception:
                pass

            # 收集非表格文本
            text = page.get_text("text")
            if text.strip():
                text_lines.append(f"--- 第 {page_num + 1} 页 ---")
                text_lines.append(text.strip())
                text_lines.append("")

        doc.close()

        # 如果没有检测到表格，创建文本内容工作表
        if table_count == 0:
            ws = wb.create_sheet(title="文本内容")
            for i, line in enumerate(text_lines):
                ws.cell(row=i + 1, column=1, value=line)

        # 如果也没有文本内容，创建一个空表
        if table_count == 0 and not text_lines:
            ws = wb.create_sheet(title="无内容")
            ws.cell(row=1, column=1, value="未在PDF中检测到表格或文本内容")

        output_path = self._make_output_path(input_path, output_dir, ".xlsx")
        wb.save(output_path)
        return output_path

    def _write_table_to_sheet(self, ws, data: list[list], format_cells: bool = True) -> None:
        """将表格数据写入工作表"""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(name="微软雅黑", bold=True, size=11)
        body_font = Font(name="微软雅黑", size=10)

        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                cell = ws.cell(row=i + 1, column=j + 1, value=str(cell_value) if cell_value else "")

                if format_cells:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

                    if i == 0:
                        cell.font = header_font
                        cell.fill = header_fill
                    else:
                        cell.font = body_font

        # 自适应列宽
        if format_cells:
            for col_idx in range(1, len(data[0]) + 1 if data else 2):
                max_length = 0
                for row_idx in range(1, len(data) + 1):
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                    # 中文字符算 2 个字符宽度
                    char_len = 0
                    for char in cell_value[:50]:  # 限制检测前 50 字符
                        char_len += 2 if ord(char) > 127 else 1
                    max_length = max(max_length, char_len)
                adjusted_width = min(max_length + 4, 60)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
